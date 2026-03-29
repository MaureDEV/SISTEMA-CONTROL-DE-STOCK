import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

DATA_FILE = "stock_data.json"

CATEGORIAS = [
    "Gaseosas y Bebidas",
    "Fiambres y Embutidos",
    "Panificados",
    "Lácteos",
    "Mercadería General",
    "Limpieza e Higiene",
    "Snacks y Golosinas",
    "Congelados",
    "Otros"
]

# ─────────────────────────── Persistencia ───────────────────────────

def cargar_datos():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_datos(datos):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)

# ─────────────────────────── Export XLSX ────────────────────────────

def exportar_xlsx(datos, ruta):
    wb = Workbook()

    # ── Hoja 1: Stock completo ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Stock Completo"

    COLOR_HEADER   = "1A3C5E"
    COLOR_SUB      = "2E86C1"
    COLOR_BAJO     = "FADBD8"
    COLOR_MEDIO    = "FEF9E7"
    COLOR_OK       = "EAFAF1"
    COLOR_ALT_ROW  = "EBF5FB"

    thin = Side(style="thin", color="CCCCCC")
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)

    logo_font   = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    sub_font    = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    normal_font = Font(name="Arial", size=10)
    bold_font   = Font(name="Arial", bold=True, size=10)

    # Título
    ws1.merge_cells("A1:J1")
    ws1["A1"] = "🛒  CONTROL DE STOCK — MINIMARKET"
    ws1["A1"].font = logo_font
    ws1["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:J2")
    ws1["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
    ws1["A2"].font = Font(name="Arial", italic=True, size=9, color="555555")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 18

    # Encabezados
    cols = ["Código de Barras", "Nombre", "Categoría", "Precio Costo",
            "Precio Venta", "Stock Actual", "Stock Mínimo", "Estado",
            "Proveedor", "Última Actualización"]
    for c, titulo in enumerate(cols, 1):
        cell = ws1.cell(row=3, column=c, value=titulo)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=COLOR_SUB)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_cell
    ws1.row_dimensions[3].height = 30

    anchos = [18, 28, 22, 14, 14, 13, 13, 12, 22, 20]
    for i, w in enumerate(anchos, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Filas de datos
    row = 4
    for idx, (codigo, p) in enumerate(sorted(datos.items(), key=lambda x: x[1].get("categoria",""))):
        stock   = p.get("stock", 0)
        minimo  = p.get("stock_minimo", 0)
        estado  = "⚠ BAJO" if stock <= minimo else ("✓ OK" if stock > minimo * 2 else "→ MEDIO")
        color_row = (COLOR_BAJO if stock <= minimo
                     else (COLOR_MEDIO if stock <= minimo * 2 else COLOR_OK))
        fill_row = PatternFill("solid", fgColor=color_row) if stock <= minimo * 2 else (
            PatternFill("solid", fgColor=COLOR_ALT_ROW) if idx % 2 == 0 else PatternFill())

        valores = [
            codigo,
            p.get("nombre", ""),
            p.get("categoria", ""),
            p.get("precio_costo", 0),
            p.get("precio_venta", 0),
            stock,
            minimo,
            estado,
            p.get("proveedor", ""),
            p.get("ultima_actualizacion", "")
        ]
        for c, val in enumerate(valores, 1):
            cell = ws1.cell(row=row, column=c, value=val)
            cell.font = normal_font
            cell.border = border_cell
            cell.fill = fill_row
            if c in (4, 5):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif c in (6, 7):
                cell.alignment = Alignment(horizontal="center")
            elif c == 8:
                cell.alignment = Alignment(horizontal="center")
                if "BAJO" in estado:
                    cell.font = Font(name="Arial", bold=True, size=10, color="C0392B")
                elif "MEDIO" in estado:
                    cell.font = Font(name="Arial", bold=True, size=10, color="D68910")
                else:
                    cell.font = Font(name="Arial", bold=True, size=10, color="1E8449")
        row += 1

    # Totales
    ws1.cell(row=row, column=1, value="TOTALES").font = bold_font
    ws1.cell(row=row, column=6,
             value=f'=SUM(F4:F{row-1})').font = bold_font
    ws1.cell(row=row, column=6).number_format = "#,##0"

    # ── Hoja 2: Lista de faltantes ──────────────────────────────────
    ws2 = wb.create_sheet("Lista de Pedidos")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "📋  LISTA DE PEDIDOS — PRODUCTOS CON STOCK BAJO"
    ws2["A1"].font = logo_font
    ws2["A1"].fill = PatternFill("solid", fgColor="922B21")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:G2")
    ws2["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
    ws2["A2"].font = Font(name="Arial", italic=True, size=9)
    ws2["A2"].alignment = Alignment(horizontal="center")

    cols2 = ["Código", "Nombre", "Categoría", "Stock Actual", "Stock Mínimo", "A Pedir", "Proveedor"]
    for c, t in enumerate(cols2, 1):
        cell = ws2.cell(row=3, column=c, value=t)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="922B21")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_cell
    ws2.row_dimensions[3].height = 28

    anchos2 = [18, 28, 22, 13, 13, 13, 22]
    for i, w in enumerate(anchos2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    faltantes = [(c, p) for c, p in datos.items() if p.get("stock", 0) <= p.get("stock_minimo", 0)]
    faltantes.sort(key=lambda x: x[1].get("categoria", ""))

    row2 = 4
    for idx, (codigo, p) in enumerate(faltantes):
        stock  = p.get("stock", 0)
        minimo = p.get("stock_minimo", 0)
        a_pedir = max(minimo * 2 - stock, 1)
        fill = PatternFill("solid", fgColor="FADBD8") if idx % 2 == 0 else PatternFill("solid", fgColor="FDEDEC")
        vals = [codigo, p.get("nombre",""), p.get("categoria",""), stock, minimo, a_pedir, p.get("proveedor","")]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row2, column=c, value=v)
            cell.font = normal_font
            cell.border = border_cell
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if c not in (2,3,7) else "left")
        row2 += 1

    if not faltantes:
        ws2.merge_cells("A4:G4")
        ws2["A4"] = "✅  ¡Todo el stock está en buen nivel!"
        ws2["A4"].font = Font(name="Arial", bold=True, size=12, color="1E8449")
        ws2["A4"].alignment = Alignment(horizontal="center")

    # ── Hoja 3: Resumen por categoría ──────────────────────────────
    ws3 = wb.create_sheet("Resumen por Categoría")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "📊  RESUMEN POR CATEGORÍA"
    ws3["A1"].font = logo_font
    ws3["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 36

    cols3 = ["Categoría", "Total Productos", "Productos en Stock Bajo", "Valor en Stock (Costo)", "Valor en Stock (Venta)"]
    for c, t in enumerate(cols3, 1):
        cell = ws3.cell(row=2, column=c, value=t)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=COLOR_SUB)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_cell
    ws3.row_dimensions[2].height = 32

    for i, w in enumerate([28, 16, 22, 22, 22], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    resumen = {}
    for codigo, p in datos.items():
        cat = p.get("categoria", "Otros")
        if cat not in resumen:
            resumen[cat] = {"total": 0, "bajos": 0, "val_costo": 0, "val_venta": 0}
        resumen[cat]["total"] += 1
        stock = p.get("stock", 0)
        if stock <= p.get("stock_minimo", 0):
            resumen[cat]["bajos"] += 1
        resumen[cat]["val_costo"] += stock * p.get("precio_costo", 0)
        resumen[cat]["val_venta"] += stock * p.get("precio_venta", 0)

    colors_cat = ["D6EAF8","D5F5E3","FEF9E7","FDEDEC","F4ECF7","FDF2E9","EAF2FF","E8F8F5","F9EBEA"]
    for idx, (cat, d) in enumerate(sorted(resumen.items())):
        fill = PatternFill("solid", fgColor=colors_cat[idx % len(colors_cat)])
        row3 = idx + 3
        vals = [cat, d["total"], d["bajos"], d["val_costo"], d["val_venta"]]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=row3, column=c, value=v)
            cell.font = normal_font
            cell.border = border_cell
            cell.fill = fill
            if c in (4, 5):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")

    wb.save(ruta)

# ─────────────────────────── GUI ────────────────────────────────────

class StockApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Control de Stock — Minimarket")
        self.root.geometry("1200x700")
        self.root.configure(bg="#1A3C5E")
        self.datos = cargar_datos()
        self.barcode_buffer = []
        self.barcode_timer = None
        self._build_ui()
        self._bind_barcode()
        self._refresh_table()

    # ── UI ──────────────────────────────────────────────────────────
    def _build_ui(self):
        # Barra superior
        top = tk.Frame(self.root, bg="#1A3C5E", pady=8)
        top.pack(fill="x")
        tk.Label(top, text="🛒  Control de Stock — Minimarket",
                 font=("Arial", 18, "bold"), bg="#1A3C5E", fg="white").pack(side="left", padx=16)

        btn_frame = tk.Frame(top, bg="#1A3C5E")
        btn_frame.pack(side="right", padx=10)
        self._btn(btn_frame, "➕ Agregar", self._abrir_agregar, "#27AE60").pack(side="left", padx=4)
        self._btn(btn_frame, "✏️ Editar",  self._editar_seleccionado, "#2980B9").pack(side="left", padx=4)
        self._btn(btn_frame, "🗑 Eliminar", self._eliminar, "#E74C3C").pack(side="left", padx=4)
        self._btn(btn_frame, "📤 Exportar XLSX", self._exportar, "#8E44AD").pack(side="left", padx=4)
        self._btn(btn_frame, "🔄 Actualizar", self._refresh_table, "#555555").pack(side="left", padx=4)

        # Filtros
        filtros = tk.Frame(self.root, bg="#2E86C1", pady=6)
        filtros.pack(fill="x")

        tk.Label(filtros, text="Buscar:", bg="#2E86C1", fg="white",
                 font=("Arial", 10)).pack(side="left", padx=(10,4))
        self.var_buscar = tk.StringVar()
        self.var_buscar.trace("w", lambda *_: self._refresh_table())
        tk.Entry(filtros, textvariable=self.var_buscar, width=22,
                 font=("Arial", 10)).pack(side="left", padx=4)

        tk.Label(filtros, text="Categoría:", bg="#2E86C1", fg="white",
                 font=("Arial", 10)).pack(side="left", padx=(12,4))
        self.var_cat = tk.StringVar(value="Todas")
        cats = ["Todas"] + CATEGORIAS
        ttk.Combobox(filtros, textvariable=self.var_cat, values=cats,
                     width=22, state="readonly").pack(side="left", padx=4)
        self.var_cat.trace("w", lambda *_: self._refresh_table())

        self.var_solo_bajos = tk.BooleanVar()
        tk.Checkbutton(filtros, text="Solo stock bajo", variable=self.var_solo_bajos,
                       bg="#2E86C1", fg="white", selectcolor="#1A3C5E",
                       activebackground="#2E86C1", activeforeground="white",
                       font=("Arial", 10),
                       command=self._refresh_table).pack(side="left", padx=12)

        # Tabla
        tabla_frame = tk.Frame(self.root, bg="#1A3C5E")
        tabla_frame.pack(fill="both", expand=True, padx=10, pady=8)

        cols = ("codigo","nombre","categoria","costo","venta","stock","minimo","estado","proveedor","actualizado")
        self.tree = ttk.Treeview(tabla_frame, columns=cols, show="headings", height=22)

        headers = ["Código","Nombre","Categoría","$ Costo","$ Venta","Stock","Mínimo","Estado","Proveedor","Actualizado"]
        widths   = [130, 200, 160, 90, 90, 70, 70, 90, 160, 140]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h, command=lambda _c=c: self._sort(_c))
            self.tree.column(c, width=w, anchor="center")
        self.tree.column("nombre", anchor="w")
        self.tree.column("proveedor", anchor="w")

        self.tree.tag_configure("bajo",  background="#FADBD8")
        self.tree.tag_configure("medio", background="#FEF9E7")
        self.tree.tag_configure("ok",    background="#EAFAF1")

        sb_v = ttk.Scrollbar(tabla_frame, orient="vertical",   command=self.tree.yview)
        sb_h = ttk.Scrollbar(tabla_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)
        sb_v.pack(side="right", fill="y")
        sb_h.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", lambda _: self._editar_seleccionado())

        # Barra de estado inferior (contiene status + firma de autor)
        footer = tk.Frame(self.root, bg="#1A3C5E")
        footer.pack(fill="x", padx=12, pady=(0, 6))

        self.lbl_status = tk.Label(footer, text="", bg="#1A3C5E", fg="#85C1E9",
                                   font=("Arial", 9), anchor="w")
        self.lbl_status.pack(side="left", fill="x", expand=True)

        # ── Firma del autor — esquina inferior derecha ──
        tk.Label(footer, text="Hecho por:  Maure Dev",
                 bg="#1A3C5E", fg="#5D8AA8",
                 font=("Arial", 8, "italic")).pack(side="right")

    def _btn(self, parent, text, cmd, color):
        return tk.Button(parent, text=text, command=cmd, bg=color, fg="white",
                         font=("Arial", 10, "bold"), relief="flat",
                         padx=10, pady=5, cursor="hand2",
                         activebackground=color, activeforeground="white")

    # ── Tabla ───────────────────────────────────────────────────────
    def _refresh_table(self):
        buscar = self.var_buscar.get().lower()
        cat_fil = self.var_cat.get()
        solo_bajos = self.var_solo_bajos.get()

        for row in self.tree.get_children():
            self.tree.delete(row)

        total = bajo = 0
        for codigo, p in self.datos.items():
            stock  = p.get("stock", 0)
            minimo = p.get("stock_minimo", 0)
            cat    = p.get("categoria", "")
            nombre = p.get("nombre", "")

            if buscar and buscar not in nombre.lower() and buscar not in codigo.lower():
                continue
            if cat_fil != "Todas" and cat != cat_fil:
                continue
            if solo_bajos and stock > minimo:
                continue

            estado = "⚠ BAJO" if stock <= minimo else ("→ MEDIO" if stock <= minimo * 2 else "✓ OK")
            tag    = "bajo"   if stock <= minimo else ("medio"  if stock <= minimo * 2 else "ok")

            self.tree.insert("", "end", iid=codigo, tags=(tag,), values=(
                codigo, nombre, cat,
                f"${p.get('precio_costo',0):,.2f}",
                f"${p.get('precio_venta',0):,.2f}",
                stock, minimo, estado,
                p.get("proveedor",""),
                p.get("ultima_actualizacion","")
            ))
            total += 1
            if stock <= minimo:
                bajo += 1

        self.lbl_status.config(
            text=f"Mostrando {total} productos  |  ⚠ Stock bajo: {bajo}  |  "
                 f"Escanear código de barras con pistola o teclado  |  "
                 f"Doble clic para editar")

    def _sort(self, col):
        rows = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        rows.sort()
        for i, (_, k) in enumerate(rows):
            self.tree.move(k, "", i)

    # ── Lector de código de barras ──────────────────────────────────
    def _bind_barcode(self):
        self.root.bind("<Key>", self._on_key)

    def _on_key(self, event):
        if event.keysym == "Return":
            codigo = "".join(self.barcode_buffer).strip()
            self.barcode_buffer.clear()
            if self.barcode_timer:
                self.root.after_cancel(self.barcode_timer)
                self.barcode_timer = None
            if codigo:
                self._procesar_escaneo(codigo)
        else:
            char = event.char
            if char and char.isprintable():
                self.barcode_buffer.append(char)
                if self.barcode_timer:
                    self.root.after_cancel(self.barcode_timer)
                self.barcode_timer = self.root.after(100, self.barcode_buffer.clear)

    def _procesar_escaneo(self, codigo):
        if codigo in self.datos:
            self._abrir_editar(codigo, modo="escaneo")
        else:
            if messagebox.askyesno("Código escaneado",
                f"Código: {codigo}\nNo existe en stock. ¿Agregar producto nuevo?"):
                self._abrir_agregar(codigo_inicial=codigo)

    # ── CRUD ────────────────────────────────────────────────────────
    def _abrir_agregar(self, codigo_inicial=""):
        self._ventana_producto(None, codigo_inicial)

    def _editar_seleccionado(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Seleccioná un producto primero.")
            return
        self._abrir_editar(sel[0])

    def _abrir_editar(self, codigo, modo="normal"):
        self._ventana_producto(codigo)

    def _ventana_producto(self, codigo=None, codigo_inicial=""):
        es_nuevo = codigo is None
        datos_p  = self.datos.get(codigo, {}) if not es_nuevo else {}

        win = tk.Toplevel(self.root)
        win.title("Nuevo producto" if es_nuevo else "Editar producto")
        win.geometry("480x560")
        win.configure(bg="#F4F6F7")
        win.grab_set()

        tk.Label(win, text="Producto", font=("Arial", 14, "bold"),
                 bg="#2E86C1", fg="white").pack(fill="x", pady=0)

        form = tk.Frame(win, bg="#F4F6F7", padx=20, pady=10)
        form.pack(fill="both", expand=True)

        campos = {}

        def fila(label, key, default="", tipo="entry", opciones=None, row=None):
            r = row if row is not None else len(campos)
            tk.Label(form, text=label, bg="#F4F6F7", font=("Arial", 10),
                     anchor="w").grid(row=r, column=0, sticky="w", pady=4)
            if tipo == "entry":
                var = tk.StringVar(value=str(datos_p.get(key, default)))
                e = tk.Entry(form, textvariable=var, font=("Arial", 10), width=32)
                e.grid(row=r, column=1, sticky="ew", padx=6, pady=4)
                campos[key] = var
            elif tipo == "combo":
                var = tk.StringVar(value=datos_p.get(key, default))
                cb = ttk.Combobox(form, textvariable=var, values=opciones,
                                  state="readonly", font=("Arial", 10), width=30)
                cb.grid(row=r, column=1, sticky="ew", padx=6, pady=4)
                campos[key] = var
            form.columnconfigure(1, weight=1)

        var_codigo = tk.StringVar(value=codigo if codigo else codigo_inicial)
        tk.Label(form, text="Código de barras", bg="#F4F6F7", font=("Arial", 10),
                 anchor="w").grid(row=0, column=0, sticky="w", pady=4)
        e_cod = tk.Entry(form, textvariable=var_codigo, font=("Arial", 10), width=32,
                         state="normal" if es_nuevo else "disabled",
                         disabledbackground="#D5D8DC")
        e_cod.grid(row=0, column=1, sticky="ew", padx=6, pady=4)

        fila("Nombre del producto", "nombre", row=1)
        fila("Categoría", "categoria", default=CATEGORIAS[0],
             tipo="combo", opciones=CATEGORIAS, row=2)
        fila("Precio de costo ($)", "precio_costo", default="0", row=3)
        fila("Precio de venta ($)", "precio_venta", default="0", row=4)
        fila("Stock actual (unidades)", "stock", default="0", row=5)
        fila("Stock mínimo (alerta)", "stock_minimo", default="5", row=6)
        fila("Proveedor", "proveedor", row=7)
        fila("Notas / observaciones", "notas", row=8)

        def guardar():
            cod = var_codigo.get().strip()
            if not cod:
                messagebox.showerror("Error", "El código de barras no puede estar vacío.", parent=win)
                return
            if es_nuevo and cod in self.datos:
                messagebox.showerror("Error", "Ya existe un producto con ese código.", parent=win)
                return
            nombre = campos["nombre"].get().strip()
            if not nombre:
                messagebox.showerror("Error", "El nombre no puede estar vacío.", parent=win)
                return
            try:
                costo  = float(campos["precio_costo"].get().replace(",","."))
                venta  = float(campos["precio_venta"].get().replace(",","."))
                stock  = int(campos["stock"].get())
                minimo = int(campos["stock_minimo"].get())
            except ValueError:
                messagebox.showerror("Error", "Revisá los campos numéricos.", parent=win)
                return

            self.datos[cod] = {
                "nombre": nombre,
                "categoria": campos["categoria"].get(),
                "precio_costo": costo,
                "precio_venta": venta,
                "stock": stock,
                "stock_minimo": minimo,
                "proveedor": campos["proveedor"].get().strip(),
                "notas": campos["notas"].get().strip(),
                "ultima_actualizacion": datetime.now().strftime("%d/%m/%Y %H:%M")
            }
            guardar_datos(self.datos)
            self._refresh_table()
            win.destroy()
            messagebox.showinfo("✅ Listo", f"Producto '{nombre}' guardado correctamente.")

        btn_bar = tk.Frame(win, bg="#F4F6F7", pady=10)
        btn_bar.pack()
        self._btn(btn_bar, "💾 Guardar", guardar, "#27AE60").pack(side="left", padx=10)
        self._btn(btn_bar, "Cancelar",  win.destroy, "#95A5A6").pack(side="left", padx=10)

    def _eliminar(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Seleccioná un producto primero.")
            return
        codigo = sel[0]
        nombre = self.datos[codigo].get("nombre", codigo)
        if messagebox.askyesno("Eliminar", f"¿Eliminás '{nombre}'?"):
            del self.datos[codigo]
            guardar_datos(self.datos)
            self._refresh_table()

    # ── Exportar ────────────────────────────────────────────────────
    def _exportar(self):
        if not self.datos:
            messagebox.showinfo("Info", "No hay productos para exportar.")
            return
        nombre_default = f"stock_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=nombre_default,
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar reporte de stock"
        )
        if ruta:
            try:
                exportar_xlsx(self.datos, ruta)
                messagebox.showinfo("✅ Exportado",
                    f"Reporte guardado en:\n{ruta}\n\n"
                    f"Hojas: Stock Completo · Lista de Pedidos · Resumen por Categoría")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo exportar:\n{e}")


# ─────────────────────────── Main ───────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    app = StockApp(root)
    root.mainloop()



#test